import pandas as pd
import json
import time
import openpyxl
import os

import chatPackages.oci_llm_api_call as llm


def write_to_excel(queries, output_file_path, content_id):
    if os.path.exists(output_file_path):
        wb = openpyxl.load_workbook(output_file_path)
        sheet = wb.active
        row_id = sheet.max_row
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Queries"
        sheet["A1"] = "query_id"
        sheet["B1"] = "query"
        sheet["C1"] = "answer"
        sheet["D1"] = "content_id"
        row_id = 1

    for query in queries:
        sheet[f"A{row_id + 1}"] = row_id
        sheet[f"B{row_id + 1}"] = query['question']
        sheet[f"C{row_id + 1}"] = query['answer']
        sheet[f"D{row_id + 1}"] = content_id
        row_id += 1

    wb.save(output_file_path)


def get_questions(file_path):
    df = pd.read_excel(file_path, engine='openpyxl')

    for index, row in df.iterrows():
        questionnaire_prompt = f'''
Your agenda is to create a FAQ questions for the information provided related to WInfoBots Solution. 
This FAQ questions will be helpful for Sales team to answer them to the clients, so create questions assuming that you heard about the topic which will be provided as an input and then answer the question based on the input provided assumingg you are an WInfoBots expert. 
Make the questions as depth as possible and elaborate answers as much as possible so my sales team have the best information to gear up for the sales discussion. 
Generate the maximum possible questions we may get based content provided.
For each question, provide a detailed and elaborate answer, in the json structure as provided below:
{{ "result":[
{{
  "question": "<question>",
  "answer": "<answer based on the content>"
}},
{{
  "question": "<question>",
  "answer": "<answer based on the content>"
}}
]
}}

Given the following text:
["{row.iloc[2]}"]
        '''
        # print(f'questionnaire_prompt: {questionnaire_prompt}')
        res_text = llm.get_llm_response(questionnaire_prompt)
        start = res_text.find('{')
        end = res_text[::-1].rfind('}')
        # print(f"start: {start}, end: {end}")
        if end:
            questions = res_text[start:-end]
        else:
            questions = res_text

        try:
            questions = json.loads(questions)  # Convert response to JSON
            write_to_excel(questions['result'], 'DownloadedFiles/SalesQueries(update).xlsx', row.iloc[0])
        except Exception as e:
            print(f"Error decoding JSON response for row {row.iloc[0]}: {e} \nstart: {start}, end: {end} \nquestions: {res_text}")

        time.sleep(5)

    print("Finished processing all questions.")


if __name__ == '__main__':
    l_file = 'DownloadedFiles/WinfoBotsSalesDocument1.xlsx'
    get_questions(l_file)
